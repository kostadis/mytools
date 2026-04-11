"""Generate tiny FlexAI-shaped workbooks for tests.

Produces two xlsx files that mirror the real FlexAI Digital Resource Companion
schemas but contain only enough data to exercise flexai_social.py:

  - tiny_flex.xlsx  (Intro + Ally + AllyResult sheets)
  - tiny_dcs.xlsx   (Intro + "Social Choice DCs" sheet)

Fixture scope:
  - One Role: Ally
  - One Size: Normal
  - Two Contexts: Passing By, Combat
  - Two Ranks: A, B

Ranges are chosen so edge-roll tests (1, 40, 41, 100 etc.) land in predictable
buckets — see test_flexai_social.py.

Run directly to materialise the fixtures next to this file:

    python3 fixtures/build_fixtures.py

Or import `build_fixtures(out_dir)` from pytest and generate them into
`tmp_path`.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


# Header columns: (RoleSize, ContextInMainWorkbookForm, Rank)
_HEADER_COLS = [
    ("Normal", "PassingBy", "A"),
    ("Normal", "PassingBy", "B"),
    ("Normal", "Combat", "A"),
    ("Normal", "Combat", "B"),
]

# Choice rows for the Ally sheet. Values are d100 ranges in the same order as
# _HEADER_COLS. "-" means the choice is unavailable on the NPC's turn for
# that (size, context, rank).
_ALLY_CHOICES = {
    "Diplomacy":    ["01-40", "01-50", "-",     "-"    ],
    "Intimidate":   ["41-50", "51-55", "01-31", "01-23"],
    "Sense Motive": ["51-70", "56-85", "32-69", "24-77"],
    "Mislead":      ["71-80", "86-90", "70-00", "78-00"],
    "Gather Info":  ["81-90", "91-95", "-",     "-"    ],
    "Subject Info": ["91-00", "96-00", "-",     "-"    ],
}

# Success-branch results — Passing By/A chosen for predictable edges.
_SUCCESS_PASSING_BY_A = {
    "TurnsHostile":     "-",
    "Leaves":           "-",
    "IgnoresYou":       "01-10",
    "Helps":            "11-30",
    "AnswerGrudgingly": "31-40",
    "AnswerOk":         "41-60",
    "AnswerWilling":    "61-80",
    "VolunteersInfo":   "81-90",
    "GrantPlutClue":    "-",
    "ChallengesYou":    "-",
    "QuestionsMotives": "91-95",
    "RevealsClue":      "96-00",
    "RedHerring":       "-",
    "Lies":             "-",
}

# Failure-branch results — Passing By/A chosen for predictable edges.
_FAILURE_PASSING_BY_A = {
    "TurnsHostile":     "01-05",
    "Leaves":           "06-20",
    "IgnoresYou":       "21-50",
    "Helps":            "-",
    "AnswerGrudgingly": "51-70",
    "AnswerOk":         "-",
    "AnswerWilling":    "-",
    "VolunteersInfo":   "-",
    "GrantPlutClue":    "-",
    "ChallengesYou":    "71-80",
    "QuestionsMotives": "81-90",
    "RevealsClue":      "-",
    "RedHerring":       "91-95",
    "Lies":             "96-00",
}

_RESULT_ROW_ORDER = [
    "TurnsHostile", "Leaves", "IgnoresYou", "Helps",
    "AnswerGrudgingly", "AnswerOk", "AnswerWilling",
    "VolunteersInfo", "GrantPlutClue", "ChallengesYou",
    "QuestionsMotives", "RevealsClue", "RedHerring", "Lies",
]


def _build_main(out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    intro = wb.create_sheet("Intro")
    intro["A1"] = "tiny fixture workbook"

    # --- Ally (NPC turn half) ---
    ally = wb.create_sheet("Ally")
    ally.cell(2, 1, "RoleSize")
    ally.cell(3, 1, "Context")
    for i, (sz, ctx, rank) in enumerate(_HEADER_COLS):
        ally.cell(2, 2 + i, sz)
        ally.cell(3, 2 + i, ctx)
        ally.cell(4, 2 + i, rank)
    for row_idx, (choice, vals) in enumerate(_ALLY_CHOICES.items()):
        ally.cell(5 + row_idx, 1, choice)
        for col_idx, val in enumerate(vals):
            ally.cell(5 + row_idx, 2 + col_idx, val)

    # --- AllyResult ---
    res = wb.create_sheet("AllyResult")
    res.cell(2, 1, "RoleSize")
    res.cell(3, 1, "Context")
    for i, (sz, ctx, rank) in enumerate(_HEADER_COLS):
        res.cell(2, 2 + i, sz)
        res.cell(3, 2 + i, ctx)
        res.cell(4, 2 + i, rank)

    # Success block: rows 5..18 (14 results). For the fixture we use the
    # same distribution for every (size, context, rank) column so all tests
    # only need one lookup.
    for row_idx, name in enumerate(_RESULT_ROW_ORDER):
        res.cell(5 + row_idx, 1, name)
        val = _SUCCESS_PASSING_BY_A[name]
        for col_idx in range(len(_HEADER_COLS)):
            res.cell(5 + row_idx, 2 + col_idx, val)

    # FAILURE label in row 21, failure block rows 22..35.
    res.cell(21, 1, "FAILURE")
    for row_idx, name in enumerate(_RESULT_ROW_ORDER):
        res.cell(22 + row_idx, 1, name)
        val = _FAILURE_PASSING_BY_A[name]
        for col_idx in range(len(_HEADER_COLS)):
            res.cell(22 + row_idx, 2 + col_idx, val)

    wb.save(out_path)


def _build_dcs(out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    intro = wb.create_sheet("Intro")
    intro["A1"] = "tiny fixture workbook"

    sheet = wb.create_sheet("Social Choice DCs")
    header = [
        "Social Role", "RoleSize", "Context", "Skill",
        "A", "B", "C", "D",
        "A", "B", "C", "D",
    ]
    for i, h in enumerate(header):
        sheet.cell(3, 1 + i, h)

    # A handful of rows: Ally/Normal/Passing By (all 6 skills) and
    # Ally/Normal/Combat (Diplomacy and Gather Info and Subject Info are
    # unavailable, the rest have DCs).
    data = [
        # Passing By — everything available.
        ("Ally", "Normal", "Passing By", "Diplomacy",
         5, 7, 10, 12,   "5 / 10", "7 / 15", "10 / 22", "12 / 28"),
        ("Ally", "Normal", "Passing By", "Intimidate",
         8, 10, 13, 15,  "8 / 13", "10 / 18", "13 / 25", "15 / 31"),
        ("Ally", "Normal", "Passing By", "Sense Motive",
         6, 8, 11, 13,   "6 / 11", "8 / 16", "11 / 23", "13 / 29"),
        ("Ally", "Normal", "Passing By", "Mislead",
         7, 10, 12, 14,  "7 / 12", "10 / 18", "12 / 24", "14 / 30"),
        ("Ally", "Normal", "Passing By", "Gather Info",
         6, 8, 11, 13,   "6 / 11", "8 / 16", "11 / 23", "13 / 29"),
        ("Ally", "Normal", "Passing By", "Subject Info",
         7, 10, 13, 15,  "7 / 12", "10 / 18", "13 / 25", "15 / 31"),
        # Combat — Diplomacy/Gather Info/Subject Info unavailable.
        ("Ally", "Normal", "Combat", "Diplomacy",
         "-", "-", "-", "-",  "-", "-", "-", "-"),
        ("Ally", "Normal", "Combat", "Intimidate",
         6, 8, 11, 13,   "6 / 11", "8 / 16", "11 / 23", "13 / 29"),
        ("Ally", "Normal", "Combat", "Sense Motive",
         6, 8, 11, 13,   "6 / 11", "8 / 16", "11 / 23", "13 / 29"),
        ("Ally", "Normal", "Combat", "Mislead",
         8, 10, 13, 15,  "8 / 13", "10 / 18", "13 / 25", "15 / 31"),
        ("Ally", "Normal", "Combat", "Gather Info",
         "-", "-", "-", "-",  "-", "-", "-", "-"),
        ("Ally", "Normal", "Combat", "Subject Info",
         "-", "-", "-", "-",  "-", "-", "-", "-"),
    ]
    for i, row in enumerate(data):
        for j, val in enumerate(row):
            sheet.cell(4 + i, 1 + j, val)

    wb.save(out_path)


def build_fixtures(out_dir: Path) -> tuple[Path, Path]:
    """Materialise both fixture workbooks into `out_dir`.

    Returns (main_path, dcs_path).
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    main_path = out_dir / "FlexAI_Social_2021_01_04.xlsx"
    dcs_path = out_dir / "FlexAI_Social_Choice_DCs_2021_01_04.xlsx"
    _build_main(main_path)
    _build_dcs(dcs_path)
    return main_path, dcs_path


if __name__ == "__main__":
    here = Path(__file__).resolve().parent
    main, dcs = build_fixtures(here)
    print(f"Wrote {main}")
    print(f"Wrote {dcs}")
