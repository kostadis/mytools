"""FlexAI Social Encounter resolver library.

Implements the FlexAI for Social Encounters rules (FlexAI Guidebook pp. 260-265)
on top of the two Digital Resource Companion spreadsheets:

  - FlexAI_Social_2021_01_04.xlsx             (NPC turn + success/failure results)
  - FlexAI_Social_Choice_DCs_2021_01_04.xlsx  (DCs per Role x Size x Context x Rank)

The library normalises all labels to canonical snake_case keys so the Excel
quirks (LullFight vs Lull, AnswerOk vs Answers, GrantPlutClue vs GrantPlotClue,
etc.) stay confined to the loader.

See RULES.md in this directory for the verbatim rulebook text.
"""

from __future__ import annotations

import random as _random
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

# ---------------------------------------------------------------------------
# Canonical vocabulary
# ---------------------------------------------------------------------------

ROLES: List[str] = ["ally", "asset", "acquaintance", "opponent", "bystander"]

SIZES: List[str] = ["normal", "minion", "elite", "solo"]

CONTEXTS: List[str] = [
    "passing_by",
    "combat",
    "lull",
    "long_rest",
    "short_rest",
    "formal_gathering",
    "informal_gathering",
]

RANKS: List[str] = ["A", "B", "C", "D"]

CHOICES: List[str] = [
    "diplomacy",
    "intimidate",
    "sense_motive",
    "mislead",
    "gather_info",
    "subject_info",
]

RESULTS: List[str] = [
    "turns_hostile",
    "leaves",
    "ignores_you",
    "helps",
    "answers_grudgingly",
    "answers",
    "answers_willingly",
    "volunteers_info",
    "grant_plot_clue",
    "challenges_you",
    "questions_motives",
    "reveals_plot_clue",
    "red_herring",
    "lies",
]

SYSTEMS = ("5e", "pf2e")

DISPLAY_NAMES: Dict[str, str] = {
    # roles
    "ally": "Ally",
    "asset": "Asset",
    "acquaintance": "Acquaintance",
    "opponent": "Opponent",
    "bystander": "Bystander",
    # sizes
    "normal": "Normal",
    "minion": "Minion",
    "elite": "Elite",
    "solo": "Solo",
    # contexts
    "passing_by": "Passing By",
    "combat": "Combat",
    "lull": "Lull in Fighting",
    "long_rest": "Long Rest",
    "short_rest": "Short Rest",
    "formal_gathering": "Formal Gathering",
    "informal_gathering": "Informal Gathering",
    # choices
    "diplomacy": "Diplomacy",
    "intimidate": "Intimidate",
    "sense_motive": "Sense Motive",
    "mislead": "Mislead",
    "gather_info": "Gather Info",
    "subject_info": "Subject Info",
    # results
    "turns_hostile": "Turns Hostile",
    "leaves": "Leaves",
    "ignores_you": "Ignores You",
    "helps": "Helps",
    "answers_grudgingly": "Answers Grudgingly",
    "answers": "Answers",
    "answers_willingly": "Answers Willingly",
    "volunteers_info": "Volunteers Info",
    "grant_plot_clue": "Can Grant Plot Clue",
    "challenges_you": "Challenges You",
    "questions_motives": "Questions Motives",
    "reveals_plot_clue": "Reveals Plot Clue",
    "red_herring": "Red Herring",
    "lies": "Lies",
}

# ---------------------------------------------------------------------------
# Label normalisation
# ---------------------------------------------------------------------------

# Maps a "compressed" form (lowercase, whitespace/underscore/hyphen stripped)
# of any known label variant to its canonical key.

_ROLE_ALIASES = {
    "ally": "ally",
    "asset": "asset",
    "acquaintance": "acquaintance",
    "opponent": "opponent",
    "bystander": "bystander",
}

_SIZE_ALIASES = {
    "normal": "normal",
    "minion": "minion",
    "elite": "elite",
    "solo": "solo",
}

_CONTEXT_ALIASES = {
    "passingby": "passing_by",
    "passing": "passing_by",  # shouldn't occur, safe-guard
    "combat": "combat",
    "lull": "lull",
    "lullfight": "lull",
    "lullinfighting": "lull",
    "longrest": "long_rest",
    "shortrest": "short_rest",
    "formalgather": "formal_gathering",
    "formalgathering": "formal_gathering",
    "informalgather": "informal_gathering",
    "informalgathering": "informal_gathering",
}

_RANK_ALIASES = {"a": "A", "b": "B", "c": "C", "d": "D"}

_CHOICE_ALIASES = {
    "diplomacy": "diplomacy",
    "intimidate": "intimidate",
    "sensemotive": "sense_motive",
    "mislead": "mislead",
    "gatherinfo": "gather_info",
    "subjectinfo": "subject_info",
}

_RESULT_ALIASES = {
    "turnshostile": "turns_hostile",
    "leaves": "leaves",
    "ignoresyou": "ignores_you",
    "helps": "helps",
    "answergrudgingly": "answers_grudgingly",
    "answersgrudgingly": "answers_grudgingly",
    "answerok": "answers",
    "answers": "answers",
    "answerwilling": "answers_willingly",
    "answerswillingly": "answers_willingly",
    "volunteersinfo": "volunteers_info",
    "grantplotclue": "grant_plot_clue",
    "grantplutclue": "grant_plot_clue",  # typo in source workbook
    "cangrantplotclue": "grant_plot_clue",
    "challengesyou": "challenges_you",
    "questionsmotives": "questions_motives",
    "revealsclue": "reveals_plot_clue",
    "revealsplotclue": "reveals_plot_clue",
    "redherring": "red_herring",
    "lies": "lies",
}


def _compress(s: object) -> str:
    if s is None:
        return ""
    return (
        str(s)
        .strip()
        .lower()
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
    )


def _lookup(aliases: Dict[str, str], value: object) -> Optional[str]:
    key = _compress(value)
    if not key:
        return None
    return aliases.get(key)


def canon_role(v: object) -> Optional[str]:
    return _lookup(_ROLE_ALIASES, v)


def canon_size(v: object) -> Optional[str]:
    return _lookup(_SIZE_ALIASES, v)


def canon_context(v: object) -> Optional[str]:
    return _lookup(_CONTEXT_ALIASES, v)


def canon_rank(v: object) -> Optional[str]:
    if v is None:
        return None
    return _RANK_ALIASES.get(str(v).strip().lower())


def canon_choice(v: object) -> Optional[str]:
    return _lookup(_CHOICE_ALIASES, v)


def canon_result(v: object) -> Optional[str]:
    return _lookup(_RESULT_ALIASES, v)


def display(key: object) -> str:
    if key is None:
        return ""
    return DISPLAY_NAMES.get(str(key), str(key))


# ---------------------------------------------------------------------------
# Value parsing
# ---------------------------------------------------------------------------

_NA_STRINGS = {"", "-", "\u2014", "\u2013", "n/a", "na"}


def parse_range(v: object) -> Optional[Tuple[int, int]]:
    """Parse a d100 range cell.

    Returns (lo, hi) inclusive, or None if the cell is blank / "-" / "n/a".

    The source workbook uses trailing "00" to mean 100 (e.g. "96-00" is the
    inclusive range 96..100) and zero-padded two-digit values ("01-40").
    """

    if v is None:
        return None
    if isinstance(v, (int, float)):
        n = int(v)
        if n == 0:
            n = 100
        return (n, n)
    s = str(v).strip()
    if s.lower() in _NA_STRINGS:
        return None
    # Some cells use an en-dash instead of a hyphen.
    for sep in ("-", "\u2013", "\u2014"):
        if sep in s:
            lo_s, _, hi_s = s.partition(sep)
            try:
                lo = int(lo_s.strip())
                hi = int(hi_s.strip())
            except ValueError:
                return None
            if lo == 0:
                lo = 100
            if hi == 0:
                hi = 100
            return (lo, hi)
    # Single number.
    try:
        n = int(s)
    except ValueError:
        return None
    if n == 0:
        n = 100
    return (n, n)


def parse_dc(v: object) -> Optional[int]:
    """Parse a plain DC cell (cols E-H of the DC sheet)."""

    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if s.lower() in _NA_STRINGS:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def parse_dc_pf2e(v: object) -> Optional[int]:
    """Parse the combined '5E / PF2e' DC cell (cols I-L of the DC sheet).

    Returns the PF2e number (the right-hand side of the slash). If the cell
    contains only a single number we return it. If the cell is blank / "-"
    we return None.
    """

    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if s.lower() in _NA_STRINGS:
        return None
    if "/" in s:
        _, _, right = s.partition("/")
        try:
            return int(right.strip())
        except ValueError:
            return None
    try:
        return int(s)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# FlexCell data model
# ---------------------------------------------------------------------------

# A FlexCell is the fully-resolved data for a (role, size, context, rank) key:
#
#   {
#     "npc_turn":        {choice_key: (lo, hi) | None, ...},
#     "success_results": {result_key: (lo, hi) | None, ...},
#     "failure_results": {result_key: (lo, hi) | None, ...},
#     "dcs_5e":          {choice_key: int | None, ...},
#     "dcs_pf2e":        {choice_key: int | None, ...},
#   }
#
# FlexTable is the full tree: table[role][size][context][rank] -> FlexCell.

FlexCell = Dict[str, Dict[str, object]]
FlexTable = Dict[str, Dict[str, Dict[str, Dict[str, FlexCell]]]]


def _new_cell() -> FlexCell:
    return {
        "npc_turn": {},
        "success_results": {},
        "failure_results": {},
        "dcs_5e": {},
        "dcs_pf2e": {},
    }


def _ensure_cell(table: FlexTable, role: str, size: str, context: str, rank: str) -> FlexCell:
    return (
        table.setdefault(role, {})
        .setdefault(size, {})
        .setdefault(context, {})
        .setdefault(rank, _new_cell())
    )


# ---------------------------------------------------------------------------
# Workbook loading
# ---------------------------------------------------------------------------

MAIN_WORKBOOK_NAME = "FlexAI_Social_2021_01_04.xlsx"
DCS_WORKBOOK_NAME = "FlexAI_Social_Choice_DCs_2021_01_04.xlsx"

_DEFAULT_DATA_DIR = Path(
    "/mnt/g/My Drive/DriveThru/Infinium Game Studios/"
    "FlexAI Digital Resource Companion (unisystem_5E_Pathfinder_P2E_OSR)"
)


def default_data_dir() -> Path:
    return _DEFAULT_DATA_DIR


def _row_headers(ws, cols: int) -> Tuple[List, List, List]:
    """Read the (RoleSize, Context, Rank) header rows (rows 2, 3, 4)."""
    sizes = [ws.cell(2, c).value for c in range(1, cols + 1)]
    contexts = [ws.cell(3, c).value for c in range(1, cols + 1)]
    ranks = [ws.cell(4, c).value for c in range(1, cols + 1)]
    return sizes, contexts, ranks


def _load_role_sheet(ws, role: str, table: FlexTable) -> None:
    """Load the NPC-turn half of a role sheet (rows 5-10 are choice rows)."""

    cols = ws.max_column
    sizes, contexts, ranks = _row_headers(ws, cols)

    for r in range(5, ws.max_row + 1):
        choice = canon_choice(ws.cell(r, 1).value)
        if not choice:
            continue
        for c in range(2, cols + 1):
            size = canon_size(sizes[c - 1])
            context = canon_context(contexts[c - 1])
            rank = canon_rank(ranks[c - 1])
            if not (size and context and rank):
                continue
            cell = _ensure_cell(table, role, size, context, rank)
            cell["npc_turn"][choice] = parse_range(ws.cell(r, c).value)


def _load_result_sheet(ws, role: str, table: FlexTable) -> None:
    """Load both result blocks.

    The sheet has rows 5-18 (success) and rows 22-35 (failure), with a
    'FAILURE' label in col A somewhere between them. We discover the blocks
    by scanning col A.
    """

    cols = ws.max_column
    sizes, contexts, ranks = _row_headers(ws, cols)

    success_rows: Dict[str, int] = {}
    failure_rows: Dict[str, int] = {}
    in_failure = False
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        if isinstance(v, str) and v.strip().upper() == "FAILURE":
            in_failure = True
            continue
        key = canon_result(v)
        if not key:
            continue
        if in_failure:
            failure_rows[key] = r
        else:
            success_rows[key] = r

    def _populate(rows_map: Dict[str, int], cell_key: str) -> None:
        for result, r in rows_map.items():
            for c in range(2, cols + 1):
                size = canon_size(sizes[c - 1])
                context = canon_context(contexts[c - 1])
                rank = canon_rank(ranks[c - 1])
                if not (size and context and rank):
                    continue
                cell = _ensure_cell(table, role, size, context, rank)
                cell[cell_key][result] = parse_range(ws.cell(r, c).value)

    _populate(success_rows, "success_results")
    _populate(failure_rows, "failure_results")


def _load_dc_sheet(ws, table: FlexTable) -> None:
    """Load the 'Social Choice DCs' sheet.

    Row 3 headers: Social Role, RoleSize, Context, Skill, A, B, C, D, A, B, C, D.
    Rows 4+ are data. Cols E-H are 5E DCs; cols I-L are combined '5E / PF2e'.
    """

    for r in range(4, ws.max_row + 1):
        role = canon_role(ws.cell(r, 1).value)
        size = canon_size(ws.cell(r, 2).value)
        context = canon_context(ws.cell(r, 3).value)
        choice = canon_choice(ws.cell(r, 4).value)
        if not (role and size and context and choice):
            continue
        for rank_idx, rank in enumerate(RANKS):
            dc_5e = parse_dc(ws.cell(r, 5 + rank_idx).value)
            dc_pf2e = parse_dc_pf2e(ws.cell(r, 9 + rank_idx).value)
            cell = _ensure_cell(table, role, size, context, rank)
            cell["dcs_5e"][choice] = dc_5e
            cell["dcs_pf2e"][choice] = dc_pf2e


def load_tables(data_dir: Path) -> FlexTable:
    """Load both FlexAI social workbooks from `data_dir`."""

    data_dir = Path(data_dir)
    main_path = data_dir / MAIN_WORKBOOK_NAME
    dcs_path = data_dir / DCS_WORKBOOK_NAME
    if not main_path.exists():
        raise FileNotFoundError(f"Main workbook not found: {main_path}")

    table: FlexTable = {}

    wb_main = openpyxl.load_workbook(main_path, data_only=True)
    for role in ROLES:
        sheet_name = display(role)  # "Ally", "Asset", ...
        result_name = sheet_name + "Result"
        if sheet_name in wb_main.sheetnames:
            _load_role_sheet(wb_main[sheet_name], role, table)
        if result_name in wb_main.sheetnames:
            _load_result_sheet(wb_main[result_name], role, table)

    if dcs_path.exists():
        wb_dcs = openpyxl.load_workbook(dcs_path, data_only=True)
        if "Social Choice DCs" in wb_dcs.sheetnames:
            _load_dc_sheet(wb_dcs["Social Choice DCs"], table)

    return table


# ---------------------------------------------------------------------------
# Query API
# ---------------------------------------------------------------------------


def list_roles() -> List[str]:
    return list(ROLES)


def list_sizes() -> List[str]:
    return list(SIZES)


def list_contexts() -> List[str]:
    return list(CONTEXTS)


def list_ranks() -> List[str]:
    return list(RANKS)


def list_choices() -> List[str]:
    return list(CHOICES)


def list_results() -> List[str]:
    return list(RESULTS)


def get_cell(
    table: FlexTable,
    role: str,
    size: str,
    context: str,
    rank: str,
) -> Optional[FlexCell]:
    try:
        return table[role][size][context][rank]
    except KeyError:
        return None


def available_choices(
    cell: Optional[FlexCell],
    system: str = "5e",
) -> List[Tuple[str, int]]:
    """Return [(choice, dc)] in canonical order for choices whose DC is not None."""

    if not cell:
        return []
    key = "dcs_5e" if system == "5e" else "dcs_pf2e"
    dcs = cell.get(key, {}) or {}
    out: List[Tuple[str, int]] = []
    for choice in CHOICES:
        dc = dcs.get(choice)
        if dc is not None:
            out.append((choice, int(dc)))
    return out


# ---------------------------------------------------------------------------
# Rolling / resolution
# ---------------------------------------------------------------------------


def roll_d100(rng: Optional[_random.Random] = None) -> int:
    r = rng if rng is not None else _random
    return r.randint(1, 100)


def _pick_bucket(
    ranges: Dict[str, Optional[Tuple[int, int]]],
    roll: int,
) -> Optional[str]:
    for key, rng in ranges.items():
        if rng is None:
            continue
        lo, hi = rng
        if lo <= roll <= hi:
            return key
    return None


def roll_npc_turn(
    cell: Optional[FlexCell],
    rng: Optional[_random.Random] = None,
) -> Tuple[int, Optional[str]]:
    """Roll d100 against the NPC-turn table. Returns (roll, choice_or_None)."""

    roll = roll_d100(rng)
    if not cell:
        return roll, None
    return roll, _pick_bucket(cell.get("npc_turn", {}), roll)


def roll_result(
    cell: Optional[FlexCell],
    success: bool,
    rng: Optional[_random.Random] = None,
) -> Tuple[int, Optional[str]]:
    """Roll d100 against the success or failure result table."""

    roll = roll_d100(rng)
    if not cell:
        return roll, None
    key = "success_results" if success else "failure_results"
    return roll, _pick_bucket(cell.get(key, {}), roll)


def attempt(
    cell: Optional[FlexCell],
    choice: str,
    pc_total: int,
    system: str = "5e",
    rng: Optional[_random.Random] = None,
) -> Dict[str, object]:
    """Resolve a single PC interaction attempt end-to-end.

    - Looks up the DC for (choice, system).
    - If the DC is None: auto-fail, no result rolled, notes explains why.
    - Otherwise compares pc_total to dc, rolls on the success/failure result
      table and returns the full breakdown.

    Does not model critical success / critical failure; callers that want to
    treat the attempt as a critical can call `roll_result` directly or pass
    the `critical=True` flag to select the immediate rather than lenient
    branch in the UI.
    """

    result: Dict[str, object] = {
        "choice": choice,
        "system": system,
        "pc_total": pc_total,
        "dc": None,
        "success": False,
        "roll": None,
        "result": None,
        "notes": None,
    }

    if not cell:
        result["notes"] = "no data for this (role, size, context, rank)"
        return result

    dc_map = cell.get("dcs_5e" if system == "5e" else "dcs_pf2e", {}) or {}
    dc = dc_map.get(choice)
    if dc is None:
        result["notes"] = "unavailable - automatic failure (no DC listed)"
        return result

    result["dc"] = int(dc)
    success = pc_total >= int(dc)
    roll, outcome = roll_result(cell, success, rng)
    result["success"] = success
    result["roll"] = roll
    result["result"] = outcome
    return result
