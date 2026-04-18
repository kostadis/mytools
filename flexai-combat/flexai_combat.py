"""FlexAI Combat Encounter resolver library.

Implements the FlexAI for Combat Encounters rules (FlexAI Guidebook pp. 12-22)
on top of the Digital Resource Companion spreadsheet:

  - FlexAI_Combat_2021_01_04.xlsx

The workbook contains one '{Role}' sheet and one '{Role}Tgt' sheet for each of
the seven Roles (Brute, Soldier, Artillery, Skirmisher, Lurker, Controller,
Leader).  Each data sheet is keyed by (RoleSize x Stance x Rank) column
headers and outcome rows (6 base outcomes, plus Minor/Major Surge and
Minor/Major Lull variants on the Outcome sheets).

This library exposes three complexity tiers:

  - "simple"   : fixed d20 Outcome and Targeting tables (from pp. 15 of the
                 Guidebook; baked in since they are not in the xlsx).
  - "full"     : d100 roll on the Role x Size x Stance x Rank FlexTable; the
                 Surge/Lull rows collapse back to their base outcome and the
                 surge tag is dropped.
  - "advanced" : same roll as "full" but the surge/lull tag is retained so the
                 GM can apply the Minor/Major Surge or Lull bonus from
                 Tables 9 and 10.

See RULES.md in this directory for the verbatim rulebook text.
"""

from __future__ import annotations

import random as _random
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

# ---------------------------------------------------------------------------
# Canonical vocabulary
# ---------------------------------------------------------------------------

ROLES: List[str] = [
    "brute",
    "soldier",
    "artillery",
    "skirmisher",
    "lurker",
    "controller",
    "leader",
]

SIZES: List[str] = ["normal", "minion", "elite", "solo"]

STANCES: List[str] = [
    "fresh",
    "ambushing",
    "unprepared",
    "bloodied",
    "cornered",
    "overwhelmed",
    "relentless",
    "mindless",
]

RANKS: List[str] = ["A", "B", "C", "D"]

OUTCOMES: List[str] = [
    "attack_main",
    "attack_secondary",
    "maneuver",
    "use_defend",
    "ability",
    "flee",
]

SURGE_LEVELS: List[Optional[str]] = [
    None,
    "minor_surge",
    "major_surge",
    "minor_lull",
    "major_lull",
]

TARGETS: List[str] = [
    "frontline",
    "rearguard",
    "closest",
    "farthest",
    "strongest",
    "weakest",
    "ranged_enemy",
    "melee_enemy",
]

TIERS = ("simple", "full", "advanced")

DISPLAY_NAMES: Dict[str, str] = {
    # roles
    "brute": "Brute",
    "soldier": "Soldier",
    "artillery": "Artillery",
    "skirmisher": "Skirmisher",
    "lurker": "Lurker",
    "controller": "Controller",
    "leader": "Leader",
    # sizes
    "normal": "Normal",
    "minion": "Minion",
    "elite": "Elite",
    "solo": "Solo",
    # stances
    "fresh": "Fresh",
    "ambushing": "Ambushing",
    "unprepared": "Unprepared",
    "bloodied": "Bloodied",
    "cornered": "Cornered",
    "overwhelmed": "Overwhelmed",
    "relentless": "Relentless",
    "mindless": "Mindless",
    # outcomes
    "attack_main": "Attack Main",
    "attack_secondary": "Attack Secondary",
    "maneuver": "Maneuver",
    "use_defend": "Use / Defend",
    "ability": "Ability",
    "flee": "Flee",
    # surge levels
    "minor_surge": "Minor Surge",
    "major_surge": "Major Surge",
    "minor_lull": "Minor Lull",
    "major_lull": "Major Lull",
    # targets
    "frontline": "Frontline",
    "rearguard": "Rearguard",
    "closest": "Closest",
    "farthest": "Farthest",
    "strongest": "Strongest",
    "weakest": "Weakest",
    "ranged_enemy": "Ranged Enemy",
    "melee_enemy": "Melee Enemy",
}


# ---------------------------------------------------------------------------
# Label normalisation
# ---------------------------------------------------------------------------

_ROLE_ALIASES = {r: r for r in ROLES}

_SIZE_ALIASES = {s: s for s in SIZES}

_STANCE_ALIASES = {s: s for s in STANCES}

_RANK_ALIASES = {"a": "A", "b": "B", "c": "C", "d": "D"}

_TARGET_ALIASES = {
    "frontline": "frontline",
    "rearguard": "rearguard",
    "closest": "closest",
    "farthest": "farthest",
    "strongest": "strongest",
    "weakest": "weakest",
    "rangedenemy": "ranged_enemy",
    "meleeenemy": "melee_enemy",
}

# Outcome-row labels are compound: "Attack Main" or "Attack Main, Minor Surge".
# _OUTCOME_ROW_ALIASES maps the compressed form to (outcome, surge_level).
_OUTCOME_ROW_ALIASES: Dict[str, Tuple[str, Optional[str]]] = {
    # base outcomes
    "attackmain": ("attack_main", None),
    "attacksecondary": ("attack_secondary", None),
    "maneuver": ("maneuver", None),
    "usedefend": ("use_defend", None),
    "use/defend": ("use_defend", None),
    "ability": ("ability", None),
    "flee": ("flee", None),
}

# Programmatically extend with the 24 surge/lull rows so the spelling table
# stays short and honest.
for _base in [
    "attackmain",
    "attacksecondary",
    "maneuver",
    "usedefend",
    "use/defend",
    "ability",
    "flee",
]:
    _canon_outcome = _OUTCOME_ROW_ALIASES[_base][0]
    for _surge_key, _surge_val in [
        ("minorsurge", "minor_surge"),
        ("majorsurge", "major_surge"),
        ("minorlull", "minor_lull"),
        ("majorlull", "major_lull"),
    ]:
        _OUTCOME_ROW_ALIASES[_base + "," + _surge_key] = (_canon_outcome, _surge_val)


def _compress(s: object) -> str:
    if s is None:
        return ""
    return (
        str(s)
        .strip()
        .lower()
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
    )


def _lookup(aliases: Dict[str, str], value: object) -> Optional[str]:
    key = _compress(value)
    if not key:
        return None
    return aliases.get(key)


def canon_role(v: object) -> Optional[str]:
    return _lookup(_ROLE_ALIASES, v)


def canon_size(v: object) -> Optional[str]:
    return _lookup(_SIZE_ALIASES, v)


def canon_stance(v: object) -> Optional[str]:
    return _lookup(_STANCE_ALIASES, v)


def canon_rank(v: object) -> Optional[str]:
    if v is None:
        return None
    return _RANK_ALIASES.get(str(v).strip().lower())


def canon_target(v: object) -> Optional[str]:
    return _lookup(_TARGET_ALIASES, v)


def canon_outcome_row(v: object) -> Optional[Tuple[str, Optional[str]]]:
    if v is None:
        return None
    return _OUTCOME_ROW_ALIASES.get(_compress(v))


def display(key: object) -> str:
    if key is None:
        return ""
    return DISPLAY_NAMES.get(str(key), str(key))


# ---------------------------------------------------------------------------
# Value parsing
# ---------------------------------------------------------------------------

_NA_STRINGS = {"", "-", "\u2014", "\u2013", "n/a", "na"}


def parse_range(v: object) -> Optional[Tuple[int, int]]:
    """Parse a d100 range cell.

    Returns (lo, hi) inclusive, or None if the cell is blank / "-" / "n/a".
    The workbook uses trailing "00" to mean 100 ("96-00" means 96..100).
    """

    if v is None:
        return None
    if isinstance(v, (int, float)):
        n = int(v)
        if n == 0:
            n = 100
        return (n, n)
    s = str(v).strip()
    if s.lower() in _NA_STRINGS:
        return None
    for sep in ("-", "\u2013", "\u2014"):
        if sep in s:
            lo_s, _, hi_s = s.partition(sep)
            try:
                lo = int(lo_s.strip())
                hi = int(hi_s.strip())
            except ValueError:
                return None
            if lo == 0:
                lo = 100
            if hi == 0:
                hi = 100
            return (lo, hi)
    try:
        n = int(s)
    except ValueError:
        return None
    if n == 0:
        n = 100
    return (n, n)


# ---------------------------------------------------------------------------
# FlexCell data model
# ---------------------------------------------------------------------------

# A FlexCell is the fully-resolved data for a (role, size, stance, rank) key:
#
#   {
#     "outcomes":  { (outcome_key, surge_level|None): (lo, hi) | None, ... },
#     "targeting": { target_key:                       (lo, hi) | None, ... },
#   }
#
# FlexTable is the full tree: table[role][size][stance][rank] -> FlexCell.

FlexCell = Dict[str, Dict]
FlexTable = Dict[str, Dict[str, Dict[str, Dict[str, FlexCell]]]]


def _new_cell() -> FlexCell:
    return {"outcomes": {}, "targeting": {}}


def _ensure_cell(
    table: FlexTable, role: str, size: str, stance: str, rank: str
) -> FlexCell:
    return (
        table.setdefault(role, {})
        .setdefault(size, {})
        .setdefault(stance, {})
        .setdefault(rank, _new_cell())
    )


# ---------------------------------------------------------------------------
# Workbook loading
# ---------------------------------------------------------------------------

WORKBOOK_NAME = "FlexAI_Combat_2021_01_04.xlsx"

_DEFAULT_DATA_DIR = Path(
    "/home/kroussos/kosta/OneDrive/Dungeons and Dragons/Tools"
)


def default_data_dir() -> Path:
    return _DEFAULT_DATA_DIR


def _row_headers(ws, cols: int) -> Tuple[List, List, List]:
    """Read the (RoleSize, Stance, Rank) header rows (rows 2, 3, 4)."""
    sizes = [ws.cell(2, c).value for c in range(1, cols + 1)]
    stances = [ws.cell(3, c).value for c in range(1, cols + 1)]
    ranks = [ws.cell(4, c).value for c in range(1, cols + 1)]
    return sizes, stances, ranks


def _load_outcome_sheet(ws, role: str, table: FlexTable) -> None:
    cols = ws.max_column
    sizes, stances, ranks = _row_headers(ws, cols)
    for r in range(5, ws.max_row + 1):
        parsed = canon_outcome_row(ws.cell(r, 1).value)
        if parsed is None:
            continue
        outcome, surge = parsed
        for c in range(2, cols + 1):
            size = canon_size(sizes[c - 1])
            stance = canon_stance(stances[c - 1])
            rank = canon_rank(ranks[c - 1])
            if not (size and stance and rank):
                continue
            cell = _ensure_cell(table, role, size, stance, rank)
            cell["outcomes"][(outcome, surge)] = parse_range(ws.cell(r, c).value)


def _load_targeting_sheet(ws, role: str, table: FlexTable) -> None:
    cols = ws.max_column
    sizes, stances, ranks = _row_headers(ws, cols)
    for r in range(5, ws.max_row + 1):
        target = canon_target(ws.cell(r, 1).value)
        if not target:
            continue
        for c in range(2, cols + 1):
            size = canon_size(sizes[c - 1])
            stance = canon_stance(stances[c - 1])
            rank = canon_rank(ranks[c - 1])
            if not (size and stance and rank):
                continue
            cell = _ensure_cell(table, role, size, stance, rank)
            cell["targeting"][target] = parse_range(ws.cell(r, c).value)


def load_tables(data_dir: Path) -> FlexTable:
    """Load the FlexAI combat workbook from `data_dir`."""

    data_dir = Path(data_dir)
    path = data_dir / WORKBOOK_NAME
    if not path.exists():
        raise FileNotFoundError(f"Combat workbook not found: {path}")

    table: FlexTable = {}
    wb = openpyxl.load_workbook(path, data_only=True)

    for role in ROLES:
        sheet_name = display(role)  # "Brute", "Soldier", ...
        tgt_name = sheet_name + "Tgt"
        if sheet_name in wb.sheetnames:
            _load_outcome_sheet(wb[sheet_name], role, table)
        if tgt_name in wb.sheetnames:
            _load_targeting_sheet(wb[tgt_name], role, table)

    return table


# ---------------------------------------------------------------------------
# Query API
# ---------------------------------------------------------------------------


def list_roles() -> List[str]:
    return list(ROLES)


def list_sizes() -> List[str]:
    return list(SIZES)


def list_stances() -> List[str]:
    return list(STANCES)


def list_ranks() -> List[str]:
    return list(RANKS)


def list_outcomes() -> List[str]:
    return list(OUTCOMES)


def list_targets() -> List[str]:
    return list(TARGETS)


def list_tiers() -> Tuple[str, ...]:
    return TIERS


def get_cell(
    table: FlexTable,
    role: str,
    size: str,
    stance: str,
    rank: str,
) -> Optional[FlexCell]:
    try:
        return table[role][size][stance][rank]
    except KeyError:
        return None


# ---------------------------------------------------------------------------
# Simple AI tables (d20; from Guidebook p. 15)
# ---------------------------------------------------------------------------

SIMPLE_OUTCOME_TABLE: List[Tuple[Tuple[int, int], str]] = [
    ((1, 12), "attack_main"),
    ((13, 14), "attack_secondary"),
    ((15, 15), "maneuver"),
    ((16, 16), "use_defend"),
    ((17, 19), "ability"),
    ((20, 20), "flee"),
]

SIMPLE_TARGET_TABLE: List[Tuple[Tuple[int, int], str]] = [
    ((1, 5), "frontline"),
    ((6, 7), "rearguard"),
    ((8, 13), "closest"),
    ((14, 14), "farthest"),
    ((15, 16), "strongest"),
    ((18, 18), "weakest"),
    ((19, 19), "ranged_enemy"),
    ((20, 20), "melee_enemy"),
]


# ---------------------------------------------------------------------------
# Rolling / resolution
# ---------------------------------------------------------------------------


def roll_d100(rng: Optional[_random.Random] = None) -> int:
    r = rng if rng is not None else _random
    return r.randint(1, 100)


def roll_d20(rng: Optional[_random.Random] = None) -> int:
    r = rng if rng is not None else _random
    return r.randint(1, 20)


def _pick_from_ordered(
    buckets: List[Tuple[Tuple[int, int], str]],
    roll: int,
) -> Optional[str]:
    for (lo, hi), key in buckets:
        if lo <= roll <= hi:
            return key
    return None


def _pick_from_dict(
    ranges: Dict,
    roll: int,
) -> Optional[object]:
    for key, rng in ranges.items():
        if rng is None:
            continue
        lo, hi = rng
        if lo <= roll <= hi:
            return key
    return None


def roll_simple_outcome(
    rng: Optional[_random.Random] = None,
) -> Tuple[int, Optional[str]]:
    roll = roll_d20(rng)
    return roll, _pick_from_ordered(SIMPLE_OUTCOME_TABLE, roll)


def roll_simple_target(
    rng: Optional[_random.Random] = None,
) -> Tuple[int, Optional[str]]:
    roll = roll_d20(rng)
    return roll, _pick_from_ordered(SIMPLE_TARGET_TABLE, roll)


def roll_full_outcome(
    cell: Optional[FlexCell],
    rng: Optional[_random.Random] = None,
) -> Tuple[int, Optional[str], Optional[str]]:
    """Roll d100 on the outcomes table.

    Returns (roll, outcome_key, surge_key).  Use Full AI by ignoring the
    surge_key return; use Advanced AI by applying the Minor/Major Surge or
    Lull bonus from Tables 9/10 in the Guidebook.
    """

    roll = roll_d100(rng)
    if not cell:
        return roll, None, None
    picked = _pick_from_dict(cell.get("outcomes", {}), roll)
    if picked is None:
        return roll, None, None
    outcome, surge = picked
    return roll, outcome, surge


def roll_targeting(
    cell: Optional[FlexCell],
    rng: Optional[_random.Random] = None,
) -> Tuple[int, Optional[str]]:
    """Roll d100 on the targeting table."""

    roll = roll_d100(rng)
    if not cell:
        return roll, None
    return roll, _pick_from_dict(cell.get("targeting", {}), roll)


def resolve_turn(
    table: FlexTable,
    role: str,
    size: str,
    stance: str,
    rank: str,
    tier: str = "full",
    rng: Optional[_random.Random] = None,
) -> Dict[str, object]:
    """Resolve a single creature's combat turn end-to-end.

    For tier="simple": roll d20 twice on the hardcoded tables; role/size/
    stance/rank are ignored.

    For tier="full" / "advanced": look up the FlexCell and roll d100 on its
    outcome + targeting tables.  "full" drops the surge tag, "advanced"
    keeps it.
    """

    if tier not in TIERS:
        raise ValueError(f"unknown tier: {tier!r}")

    result: Dict[str, object] = {
        "tier": tier,
        "role": role,
        "size": size,
        "stance": stance,
        "rank": rank,
        "outcome": None,
        "outcome_roll": None,
        "surge": None,
        "target": None,
        "target_roll": None,
        "notes": None,
    }

    if tier == "simple":
        oroll, outcome = roll_simple_outcome(rng)
        troll, target = roll_simple_target(rng)
        result["outcome_roll"] = oroll
        result["outcome"] = outcome
        result["target_roll"] = troll
        result["target"] = target
        return result

    cell = get_cell(table, role, size, stance, rank)
    if cell is None:
        result["notes"] = "no data for this (role, size, stance, rank)"
        return result

    oroll, outcome, surge = roll_full_outcome(cell, rng)
    troll, target = roll_targeting(cell, rng)
    result["outcome_roll"] = oroll
    result["outcome"] = outcome
    if tier == "advanced":
        result["surge"] = surge
    result["target_roll"] = troll
    result["target"] = target
    return result
