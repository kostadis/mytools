"""Generate a tiny FlexAI-Combat-shaped workbook for tests.

Produces one xlsx file that mirrors the real FlexAI_Combat_2021_01_04.xlsx
schema but contains only enough data to exercise flexai_combat.py:

    FlexAI_Combat_2021_01_04.xlsx   (Intro + Brute + BruteTgt sheets)

Fixture scope:
  - One Role: Brute
  - One Size: Normal
  - Two Stances: Fresh, Ambushing
  - Two Ranks: A, B

Ranges are chosen so edge-roll tests land in predictable buckets — see
test_flexai_combat.py.

Run directly to materialise the fixture next to this file:

    python3 fixtures/build_fixtures.py

Or import `build_fixtures(out_dir)` from pytest and generate it into
`tmp_path`.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


# Header columns: (RoleSize, Stance, Rank)
_HEADER_COLS = [
    ("Normal", "Fresh", "A"),
    ("Normal", "Fresh", "B"),
    ("Normal", "Ambushing", "A"),
    ("Normal", "Ambushing", "B"),
]

# Outcome rows on the Brute sheet. Values are d100 ranges in the same order
# as _HEADER_COLS. "-" means the bucket is empty for that (size, stance,
# rank).  Ranges partition 1..100 per column for deterministic roll tests.
# Column A (Normal, Fresh, A) layout:
#     attack_main        01-30
#     attack_secondary   31-50
#     maneuver           51-60
#     use_defend         61-65
#     ability            66-80
#     flee               -
#     attack_main MinS   81-85
#     ...                (remaining surge/lull buckets fill up to 100)
_BRUTE_ROWS = [
    # (row label,                    Normal/Fresh/A, Normal/Fresh/B,
    #                                Normal/Ambushing/A, Normal/Ambushing/B)
    ("Attack Main",                  ["01-30", "01-40", "01-60", "01-70"]),
    ("Attack Secondary",             ["31-50", "41-60", "-",     "-"    ]),
    ("Maneuver",                     ["51-60", "61-70", "61-70", "71-80"]),
    ("Use/Defend",                   ["61-65", "71-75", "-",     "-"    ]),
    ("Ability",                      ["66-80", "76-90", "71-90", "81-95"]),
    ("Flee",                         ["-",     "-",     "-",     "-"    ]),
    ("Attack Main, Minor Surge",     ["81-85", "91-92", "91-93", "96-97"]),
    ("Attack Secondary, Minor Surge",["86-88", "93",    "-",     "-"    ]),
    ("Maneuver, Minor Surge",        ["89-90", "94",    "94",    "98"   ]),
    ("Use/Defend, Minor Surge",      ["91",    "95",    "-",     "-"    ]),
    ("Ability, Minor Surge",         ["92-93", "96",    "95",    "99"   ]),
    ("Flee, Minor Surge",            ["-",     "-",     "-",     "-"    ]),
    ("Attack Main, Major Surge",     ["94",    "97",    "96",    "-"    ]),
    ("Attack Secondary, Major Surge",["-",     "-",     "-",     "-"    ]),
    ("Maneuver, Major Surge",        ["95",    "-",     "-",     "-"    ]),
    ("Use/Defend, Major Surge",      ["-",     "-",     "-",     "-"    ]),
    ("Ability, Major Surge",         ["96",    "98",    "97",    "-"    ]),
    ("Flee, Major Surge",            ["-",     "-",     "-",     "-"    ]),
    ("Attack Main, Minor Lull",      ["97",    "-",     "98",    "-"    ]),
    ("Attack Secondary, Minor Lull", ["-",     "-",     "-",     "-"    ]),
    ("Maneuver, Minor Lull",         ["-",     "-",     "-",     "-"    ]),
    ("Use/Defend, Minor Lull",       ["-",     "-",     "-",     "-"    ]),
    ("Ability, Minor Lull",          ["98",    "99",    "99",    "-"    ]),
    ("Flee, Minor Lull",             ["-",     "-",     "-",     "-"    ]),
    ("Attack Main, Major Lull",      ["99",    "00",    "-",     "00"   ]),
    ("Attack Secondary, Major Lull", ["-",     "-",     "-",     "-"    ]),
    ("Maneuver, Major Lull",         ["-",     "-",     "-",     "-"    ]),
    ("Use/Defend, Major Lull",       ["-",     "-",     "-",     "-"    ]),
    ("Ability, Major Lull",          ["00",    "-",     "00",    "-"    ]),
    ("Flee, Major Lull",             ["-",     "-",     "-",     "-"    ]),
]

# Targeting rows on the BruteTgt sheet.  Each column partitions 1..100
# cleanly across the 8 targets so roll tests are deterministic.
_BRUTE_TGT_ROWS = [
    ("Frontline",    ["01-20", "01-25", "01-30", "01-25"]),
    ("Rearguard",    ["21-30", "26-35", "31-40", "26-35"]),
    ("Closest",      ["31-55", "36-55", "41-60", "36-60"]),
    ("Farthest",     ["56-65", "56-65", "61-70", "61-70"]),
    ("Strongest",    ["66-80", "66-80", "71-85", "71-80"]),
    ("Weakest",      ["81-90", "81-90", "86-90", "81-90"]),
    ("Ranged Enemy", ["91-95", "91-95", "91-95", "91-95"]),
    ("Melee Enemy",  ["96-00", "96-00", "96-00", "96-00"]),
]


def _write_headers(ws) -> None:
    ws.cell(2, 1, "RoleSize")
    ws.cell(3, 1, "Stance")
    for i, (sz, st, rk) in enumerate(_HEADER_COLS):
        ws.cell(2, 2 + i, sz)
        ws.cell(3, 2 + i, st)
        ws.cell(4, 2 + i, rk)


def _build(out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    intro = wb.create_sheet("Intro")
    intro["A1"] = "tiny fixture workbook"

    brute = wb.create_sheet("Brute")
    _write_headers(brute)
    for row_idx, (label, vals) in enumerate(_BRUTE_ROWS):
        brute.cell(5 + row_idx, 1, label)
        for col_idx, val in enumerate(vals):
            brute.cell(5 + row_idx, 2 + col_idx, val)

    tgt = wb.create_sheet("BruteTgt")
    _write_headers(tgt)
    for row_idx, (label, vals) in enumerate(_BRUTE_TGT_ROWS):
        tgt.cell(5 + row_idx, 1, label)
        for col_idx, val in enumerate(vals):
            tgt.cell(5 + row_idx, 2 + col_idx, val)

    wb.save(out_path)


def build_fixtures(out_dir: Path) -> Path:
    """Materialise the combat fixture workbook into `out_dir`."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "FlexAI_Combat_2021_01_04.xlsx"
    _build(path)
    return path


if __name__ == "__main__":
    here = Path(__file__).resolve().parent
    path = build_fixtures(here)
    print(f"Wrote {path}")
